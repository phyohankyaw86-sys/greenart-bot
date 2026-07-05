"""
GreenArt Beef Jerky — Excel → Supabase (PostgreSQL) Importer
================================================================
ဒီ script ဟာ GreenArt_BeefJerky_MASTER_Workbook.xlsx ထဲက Sales_Data
နှင့် Production_Cost (Raw Material + Packaging + Labor + Utilities
အကြောင်းအလေးလေးအားလုံး) data များကို Supabase (PostgreSQL) ထဲ
တိုက်ရိုက် insert လုပ်ပေးပါတယ်။

မှတ်ချက်: ရှေးက import_excel.py ဟာ sqlite3 (local greenart.db) ကို
ချိတ်ဆက်ထားခဲ့ပြီး၊ Flask dashboard က Supabase (psycopg2) ကို
ချိတ်ဆက်နေတဲ့အတွက် data က dashboard မှာ "0" ပြနေခဲ့ပါတယ်။
ဒီ script က database.py ထဲက get_conn() logic အတိုင်း
DATABASE_URL environment variable ကိုသုံးပြီး Supabase ကို
တိုက်ရိုက်ချိတ်ဆက်ပါတယ်။

အသုံးပြုပုံ:
    1. .env file ထဲမှာ DATABASE_URL ရှိကြောင်းသေချာပါ (Supabase connection string)
    2. pip install pandas openpyxl psycopg2-binary python-dotenv
    3. python import_excel_to_supabase.py
"""

import os
import sys
import pandas as pd
import psycopg2
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

FILE = "GreenArt_BeefJerky_MASTER_Workbook.xlsx"


def get_conn():
    url = os.environ.get("DATABASE_URL")
    if not url:
        print("❌ DATABASE_URL .env (\u101e\u102d\u102f\u1037\u1019\u101f\u102f\u1010\u103a Railway Variables) \u1011\u1032\u1019\u101e\u1031\u1037\u1015\u102b! \u1011\u100a\u103a\u1015\u1031\u1038\u1015\u102b\u101c\u102f\u1004\u103a\u1006\u1031\u102c\u1004\u103a\u1015\u102b\u104a script \u1014\u101a\u103a\u1010\u1004\u103a\u1015\u101a\u103a\u1010\u102d\u102f\u1004\u103a\u1015\u102b\u1011\u103e\u1004\u103a\u1011\u100a\u103a\u1011\u102c\u1038\u1015\u102b.")
        sys.exit(1)
    return psycopg2.connect(url)


def import_sales(conn):
    print("=== Sales_Data import \u1005\u1010\u1004\u103a\u1015\u102b ===")
    df = pd.read_excel(FILE, sheet_name="Sales_Data", header=None)
    header_row = None
    for i, row in df.iterrows():
        if "Date" in str(row.values) and "Customer" in str(row.values):
            header_row = i
            break
    if header_row is None:
        print("❌ Sales_Data header \u101b\u103e\u102d\u1015\u102b")
        return 0

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df = df.dropna(subset=["Date"])

    c = conn.cursor()
    count = 0
    skipped = 0
    for _, row in df.iterrows():
        try:
            date_val = row.get("Date", "")
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]

            customer = str(row.get("Customer Name", "") or "")
            channel = str(row.get("Sale Channel", "") or "")
            item = str(row.get("Items", "") or "")
            quantity = int(float(row.get("Quantity (pcs)", 0) or 0))
            unit_price = int(float(row.get("Unit Price (MMK)", 0) or 0))
            total = int(float(row.get("Total Revenue (MMK)", 0) or 0))

            c.execute(
                """
                INSERT INTO sales (date, customer, channel, item, quantity, unit_price, total, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (date_str, customer, channel, item, quantity, unit_price, total, datetime.now().isoformat()),
            )
            count += 1
        except Exception as e:
            skipped += 1
            print(f"  \u26a0\ufe0f row skip: {e}")
    conn.commit()
    print(f"✅ Sales: {count} rows imported, {skipped} skipped\n")
    return count


def import_production_category(conn, start_row, end_row, sheet_ws, category_label):
    """Raw Material / Packaging blocks share the same Item/Qty/Unit/Cost columns."""
    c = conn.cursor()
    count = 0
    for r in range(start_row, end_row):
        no_val = sheet_ws.cell(row=r, column=1).value
        if not isinstance(no_val, (int, float)):
            continue
        date_val = sheet_ws.cell(row=r, column=2).value
        item = sheet_ws.cell(row=r, column=3).value
        qty_used = sheet_ws.cell(row=r, column=4).value
        unit = sheet_ws.cell(row=r, column=5).value
        unit_cost = sheet_ws.cell(row=r, column=6).value
        total_cost = sheet_ws.cell(row=r, column=7).value

        if date_val is None:
            continue
        try:
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]

            c.execute(
                """
                INSERT INTO production (date, item, qty_used, unit, unit_cost, total_cost)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (
                    date_str,
                    f"[{category_label}] {item}",
                    float(qty_used or 0),
                    str(unit or ""),
                    int(float(unit_cost or 0)),
                    int(float(total_cost or 0)),
                ),
            )
            count += 1
        except Exception as e:
            print(f"  \u26a0\ufe0f {category_label} row {r} skip: {e}")
    conn.commit()
    return count


def import_production(conn):
    print("=== Production_Cost import \u1005\u1010\u1004\u103a\u1015\u102b ===")
    import openpyxl
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb["Production_Cost"]

    total = 0
    # Raw Material Cost: header row 5, data rows 6-80
    total += import_production_category(conn, 6, 81, ws, "Raw Material")
    # Packaging Cost: header row 84, data rows 85-94
    total += import_production_category(conn, 85, 95, ws, "Packaging")

    print(f"✅ Production (Raw Material + Packaging): {total} rows imported")
    print("  \u2139\ufe0f  Labor \u1014\u100a\u103a\u1037 Utilities \u101e\u102d\u102f\u1037 monthly \u1006\u102d\u102f\u1010\u102c\u1019\u101a\u103a\u1019\u101e\u102f\u1014\u103e\u1004\u103a\u1037\u1015\u102b\u1010\u1031\u102c\u1037 (per-transaction \u1019\u101f\u102f\u1010\u103a) \u1010\u100a\u103a\u1006\u102d\u102f\u1004\u103a\u1038 assumptions \u1011\u1032\u1015\u102b import \u1014\u1031\u102c\u1019\u101a\u103a.\n")
    return total


def import_assumptions(conn):
    print("=== Assumptions import \u1005\u1010\u1004\u103a\u1015\u102b ===")
    import openpyxl
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb["Assumptions"]

    c = conn.cursor()
    count = 0
    for r in range(1, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if key is None or val is None:
            continue
        key = str(key).strip()
        if not isinstance(val, (int, float)):
            continue
        try:
            c.execute(
                "INSERT INTO assumptions (key, value) VALUES (%s, %s) "
                "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
                (key, float(val)),
            )
            count += 1
        except Exception as e:
            print(f"  \u26a0\ufe0f assumptions row {r} skip: {e}")
    conn.commit()
    print(f"✅ Assumptions: {count} rows imported\n")
    return count


def check_existing_data(conn):
    """Import မလုပ်ခင် duplicate ဖြစ်နိုင်ချေကို စစ်ပါ"""
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM sales")
    sales_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM production")
    prod_count = c.fetchone()[0]
    return sales_count, prod_count


if __name__ == "__main__":
    conn = get_conn()

    existing_sales, existing_prod = check_existing_data(conn)
    if existing_sales > 0 or existing_prod > 0:
        print(f"⚠️  Supabase \u1011\u1032 \u101b\u103e\u102d\u1015\u103c\u100a\u103a\u1038\u101e\u102c\u1038 data \u101b\u103e\u102d\u1015\u102b\u1010\u101a\u103a: sales={existing_sales}, production={existing_prod}")
        answer = input("Import \u1006\u1000\u103a\u101c\u102f\u1015\u103a\u1019\u101c\u102c\u1038? Duplicate \u1018\u103c\u1005\u103a\u1014\u102d\u102f\u1004\u103a\u1015\u102b\u1010\u101a\u103a! (yes/no): ")
        if answer.strip().lower() != "yes":
            print("Import \u1019\u101c\u102f\u1015\u103a\u1015\u102b\u1018\u100a\u103a\u104a \u1011\u1031\u102c\u1000\u103a\u1015\u102b\u1010\u101a\u103a.")
            conn.close()
            sys.exit(0)

    import_sales(conn)
    import_production(conn)
    import_assumptions(conn)

    conn.close()
    print("🎉 Import အကုန်ပြီးပြီ! Dashboard ကို refresh လုပ်ကြည့်ပါ.")