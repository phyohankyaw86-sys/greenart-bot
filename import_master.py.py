"""
GreenArt — MASTER Workbook → Supabase Importer
================================================
Imports: Sales_Data, Production_Cost, Yield_Tracking,
         Expense_Ledger, Inventory, Assumptions
from GreenArt_BeefJerky_MASTER_Workbook.xlsx

Usage:
    python import_master.py
    python import_master.py --clear   # wipe tables first (re-import)
"""

import os, sys, argparse
import openpyxl, psycopg2
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()
FILE = "GreenArt_BeefJerky_MASTER_Workbook.xlsx"


def get_conn():
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("❌ DATABASE_URL not set")
    return psycopg2.connect(url)


def to_date(val):
    if val is None: return None
    if hasattr(val, "strftime"): return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    try:
        datetime.strptime(s, "%Y-%m-%d"); return s
    except: return None


def to_int(val, default=0):
    try: return int(float(str(val).replace(",", ""))) if val not in (None,"","nan") else default
    except: return default


def to_float(val, default=0.0):
    try: return float(str(val).replace(",", "")) if val not in (None,"","nan") else default
    except: return default


def find_header_row(ws, required_cols, max_scan=15):
    """Find the row index where all required_cols appear."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        row_str = [str(c) for c in row if c is not None]
        if all(any(req in cell for cell in row_str) for req in required_cols):
            return i  # 0-indexed
    return None


# ── ASSUMPTIONS ───────────────────────────────────────────────────────────────
def import_assumptions(conn, wb):
    ws = wb["Assumptions"]
    c = conn.cursor()
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        key, val = row[0], row[1]
        if key is None or str(key).startswith("🥩") or str(key).startswith("Legend"):
            continue
        key = str(key).strip()
        if not key or key in ("nan","None"): continue
        if isinstance(val, (int, float)):
            c.execute("""
                INSERT INTO assumptions (key, value)
                VALUES (%s,%s)
                ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value, updated_at=NOW()
            """, (key, val))
            count += 1
    conn.commit()
    print(f"  ✅ Assumptions: {count} rows")


# ── SALES_DATA ────────────────────────────────────────────────────────────────
def import_sales(conn, wb):
    ws = wb["Sales_Data"]
    # Header: No | Date | Customer Name | Sale Channel | Delivery Channel |
    #         Delivery Charge | Items | Quantity | Unit Price | Subtotal |
    #         Delivery Income | Total Revenue
    hi = find_header_row(ws, ["Date", "Customer"], max_scan=10)
    if hi is None:
        print("  ⚠️ Sales_Data header not found"); return 0

    rows = list(ws.iter_rows(min_row=hi+2, values_only=True))
    c = conn.cursor()
    count = skipped = 0
    for row in rows:
        if not row or row[0] is None: continue
        try:
            date = to_date(row[1])
            if not date: continue
            c.execute("""
                INSERT INTO sales
                  (date,customer,channel,delivery_ch,delivery_fee,
                   item,quantity,unit_price,subtotal,delivery_inc,total_revenue)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                date,
                str(row[2] or "").strip(),
                str(row[3] or "").strip(),
                str(row[4] or "Direct").strip(),
                to_int(row[5]),
                str(row[6] or "").strip(),
                to_int(row[7]),
                to_int(row[8]),
                to_int(row[9]),
                to_int(row[10]),
                to_int(row[11]),
            ))
            count += 1
        except Exception as e:
            skipped += 1
    conn.commit()
    print(f"  ✅ Sales: {count} imported, {skipped} skipped")
    return count


# ── PRODUCTION_COST ───────────────────────────────────────────────────────────
def import_production_cost(conn, wb):
    ws = wb["Production_Cost"]
    c = conn.cursor()
    count = skipped = 0

    # Raw Material block: header around row 5, data rows 6-80
    # Packaging block: header around row 84, data rows 85-94
    # Labor/Utilities: monthly totals (no line items), handled via Assumptions

    blocks = [
        (6, 81, "raw_material"),
        (85, 100, "packaging"),
    ]
    for start, end, category in blocks:
        for r in range(start, end):
            row = [ws.cell(r, col).value for col in range(1, 8)]
            if not isinstance(row[0], (int, float)): continue
            date = to_date(row[1])
            if not date: continue
            try:
                c.execute("""
                    INSERT INTO production_cost
                      (date,batch_ref,category,item,qty_used,unit,unit_cost,total_cost)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    date, None, category,
                    str(row[2] or "").strip(),
                    to_float(row[3]),
                    str(row[4] or "").strip(),
                    to_int(row[5]),
                    to_int(row[6]),
                ))
                count += 1
            except Exception as e:
                skipped += 1
    conn.commit()
    print(f"  ✅ Production_Cost: {count} rows, {skipped} skipped")


# ── YIELD_TRACKING ────────────────────────────────────────────────────────────
def import_yield(conn, wb):
    ws = wb["Yield_Tracking"]
    # Header: No | Date | Batch No | Raw Beef Input (kg) | Dried Output (kg) |
    #         Actual Yield % | Target Yield % | Remark
    hi = find_header_row(ws, ["Date","Batch"], max_scan=10)
    if hi is None:
        print("  ⚠️ Yield_Tracking header not found"); return

    rows = list(ws.iter_rows(min_row=hi+2, values_only=True))
    c = conn.cursor()
    count = 0
    for row in rows:
        if not row or row[0] is None: continue
        date = to_date(row[1])
        raw_kg = to_float(row[3])
        dry_kg = to_float(row[4])
        if not date or raw_kg == 0: continue
        actual = round(dry_kg / raw_kg, 4) if raw_kg > 0 else 0
        target = to_float(row[6], 0.6)
        try:
            c.execute("""
                INSERT INTO yield_tracking
                  (date,batch_no,raw_beef_kg,dried_output_kg,
                   actual_yield_pct,target_yield_pct,remark)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                date, str(row[2] or "").strip(),
                raw_kg, dry_kg, actual, target,
                str(row[7] or "").strip(),
            ))
            count += 1
        except: pass
    conn.commit()
    print(f"  ✅ Yield_Tracking: {count} rows")


# ── EXPENSE_LEDGER ────────────────────────────────────────────────────────────
def import_expenses(conn, wb):
    ws = wb["Expense_Ledger"]
    hi = find_header_row(ws, ["Date","Category","Amount"], max_scan=10)
    if hi is None:
        print("  ⚠️ Expense_Ledger header not found"); return

    rows = list(ws.iter_rows(min_row=hi+2, values_only=True))
    c = conn.cursor()
    count = 0
    for row in rows:
        if not row or row[0] is None: continue
        date = to_date(row[1])
        if not date: continue
        try:
            c.execute("""
                INSERT INTO expenses
                  (date,description,category,amount,payment_method,remark)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (
                date,
                str(row[2] or "").strip(),
                str(row[3] or "Other").strip(),
                to_int(row[4]),
                str(row[5] or "Cash").strip(),
                str(row[6] or "").strip(),
            ))
            count += 1
        except: pass
    conn.commit()
    print(f"  ✅ Expenses: {count} rows")


# ── INVENTORY ────────────────────────────────────────────────────────────────
def import_inventory(conn, wb):
    ws = wb["Inventory"]
    # Multi-block sheet: Raw Beef, Packaging, Finished Goods
    # We scan for item names and their opening stock
    c = conn.cursor()
    count = 0

    # Category markers in column A
    current_cat = "raw_material"
    item_name = None
    col_map = None  # maps col index → field name once header found

    for row in ws.iter_rows(min_row=1, values_only=True):
        a = str(row[0] or "").strip()
        if "Raw Material" in a: current_cat = "raw_material"; continue
        if "Packaging" in a and "Inventory" in a: current_cat = "packaging"; continue
        if "Finished" in a: current_cat = "finished_goods"; continue

        # Item name row (just text in col A, no number in col A)
        if a and row[1] is None and row[3] is None and not a.startswith("No"):
            item_name = a; col_map = None; continue

        # Header row
        if "No" in a and row[1] is not None and "Date" in str(row[1]):
            col_map = {}
            for i, h in enumerate(row):
                h = str(h or "").strip()
                if "Opening" in h: col_map["open"] = i
                elif "Stock In" in h: col_map["in"] = i
                elif "Stock Out" in h: col_map["out"] = i
                elif "Closing Stock" in h: col_map["close"] = i
                elif "Unit Cost" in h: col_map["cost"] = i
                elif "Date" in h: col_map["date"] = i
                elif "Batch" in h: col_map["batch"] = i
            continue

        # Data row
        if col_map and isinstance(row[0], (int, float)) and item_name:
            date = to_date(row[col_map.get("date")])
            close = to_float(row[col_map.get("close", 6)])
            cost = to_int(row[col_map.get("cost", 7)])
            if not date: continue
            try:
                # Upsert into inventory_current for latest close
                c.execute("""
                    INSERT INTO inventory_current (item, category, quantity, unit, unit_cost)
                    VALUES (%s,%s,%s,%s,%s)
                    ON CONFLICT (item) DO UPDATE SET
                      quantity=EXCLUDED.quantity, unit_cost=EXCLUDED.unit_cost,
                      updated_at=NOW()
                """, (item_name, current_cat, close, "kg" if current_cat=="raw_material" else "pc", cost))
                count += 1
            except: pass

    conn.commit()
    print(f"  ✅ Inventory: {count} current snapshots upserted")


# ── CLEAR TABLES ─────────────────────────────────────────────────────────────
def clear_tables(conn):
    c = conn.cursor()
    tables = ["yield_tracking","sales","production_cost","production_batches",
              "expenses","inventory_log","inventory_current","cash_in","cash_out","assumptions"]
    for t in tables:
        c.execute(f"DELETE FROM {t}")
    conn.commit()
    print("🗑️  All tables cleared")


# ── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--clear", action="store_true", help="Clear tables before import")
    args = parser.parse_args()

    print(f"\n📂 Loading {FILE} ...")
    try:
        wb = openpyxl.load_workbook(FILE, data_only=True)
    except FileNotFoundError:
        sys.exit(f"❌ {FILE} not found in current directory")

    conn = get_conn()
    print(f"✅ Connected to database\n")

    if args.clear:
        clear_tables(conn)

    print("⏳ Importing...")
    import_assumptions(conn, wb)
    import_sales(conn, wb)
    import_production_cost(conn, wb)
    import_yield(conn, wb)
    import_expenses(conn, wb)
    import_inventory(conn, wb)

    conn.close()
    print("\n🎉 Import complete! Dashboard ကို refresh လုပ်ကြည့်ပါ")